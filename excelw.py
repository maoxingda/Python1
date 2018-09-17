from openpyxl import Workbook

if __name__ == '__main__':
    wb = Workbook()

    ws = wb.active

    ws.title = u'计科1101班'

    ws.cell(1, 1, u'学号')
    ws.cell(1, 2, u'姓名')
    ws.cell(1, 3, u'性别')
    ws.cell(1, 4, u'籍贯')

    ws.cell(2, 1, u'110101')
    ws.cell(2, 2, u'毛兴达')
    ws.cell(2, 3, u'男')
    ws.cell(2, 4, u'贵州省兴义市')

    wb.save(u'wb.xlsx')
