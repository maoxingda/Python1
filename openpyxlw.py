from openpyxl import Workbook
from openpyxl.compat import range


def create():
    wb = Workbook()
    fname = u'persons.xlsx'
    ws_person = wb.active
    ws_person.title = u'计科1101班'

    for row in range(3):
        ws_person.append(range(600))

    ws_grade = wb.create_sheet(title = u'成绩')

    ws_grade[u'F5'] = 3.14

    ws_course = wb.create_sheet(title = u'课程')

    for row in range(1, 3):
        for col in range(1, 3):
            _ = ws_course.cell(column = col, row = row, value = u'{0}')

    print(ws_course[u'AA10'].value)

    wb.save(filename = fname)


if __name__ == '__main__':
    create()
